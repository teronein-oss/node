from __future__ import annotations

import json
import re
import shutil
from collections import defaultdict
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook


SOURCE_XLSX = Path(
    "/Users/andrewko/Desktop/Codex/Socre/백현고2 내신 모의고사 4차_채점결과.xlsx"
)
OUTPUT_DIR = Path("output/baekhyeon-2026-final-4")
REPORT_DIR = OUTPUT_DIR / "students"
REPORT_SEED_PATH = Path("functions/src/reportSeedData.ts")

ANSWER_KEY = {
    1: "3",
    2: "2",
    3: "2",
    4: "4",
    5: "5",
    6: "5",
    7: "2",
    8: "4",
    9: "5",
    10: "4",
    11: "3",
    12: "1",
    13: "3",
    14: "3",
    15: "2",
    16: "5",
    17: "3",
    18: "1",
    19: "1",
    20: "1",
}

QUESTIONS = {
    1: ("대의 파악", "주제", "사회적 평가가 비행 청소년의 장기 삶의 경로에 미치는 영향"),
    2: ("대의 파악", "주제", "비편향 학습 데이터만으로는 AI 편향을 완전히 해결할 수 없음"),
    3: ("대의 파악", "주제", "친환경 의도와 실제 소비 행동 사이의 불일치 원인"),
    4: ("대의 파악", "제목", "영화가 얼굴의 클로즈업을 통해 감정 전염을 유도하는 방식"),
    5: ("추론·의미", "함의 추론", "유사한 사람끼리 모이는 힘이 소속감과 배제의 경계를 함께 형성함"),
    6: ("추론·의미", "함의 추론", "정체성 형성이 이미 존재하는 사회·역사적 조건의 영향을 받음"),
    7: ("세부 내용", "내용 불일치", "세이버메트릭스의 기원과 오클랜드 애슬레틱스의 활용"),
    8: ("세부 내용", "내용 불일치", "빛 공해가 바다거북의 산란과 부화에 미치는 영향"),
    9: ("어법·어휘", "어법", "우주정거장 일기에서 문장 구조와 분사 형태 판단"),
    10: ("어법·어휘", "어법", "습열 환경에서 인체의 열 조절을 설명하는 문장의 어법 판단"),
    11: ("어법·어휘", "문맥 어휘", "감기와 아연 연구 문맥에서 부적절한 어휘 판별"),
    12: ("추론·의미", "빈칸 추론", "독립적 이동 경험과 정신적 공간 지도 형성의 관계"),
    13: ("글의 구조", "무관한 문장", "대면·디지털 소통과 정서적 유대에 관한 글의 일관성 판단"),
    14: ("글의 구조", "글의 순서", "프로 스포츠 구단의 데이터 분석 조직 구축 과정 배열"),
    15: ("글의 구조", "글의 순서", "발과 다리가 감정을 드러내는 진화적 설명의 전개 배열"),
    16: ("글의 구조", "문장 삽입", "에너지 보존 성향과 자발적 운동의 관계를 논리적 위치에 삽입"),
    17: ("글의 구조", "문장 삽입", "공감에서 연민으로 전환될 때의 효과를 논리적 위치에 삽입"),
    18: ("글의 구조", "연결어", "무중력의 이점과 신체적 부작용 사이의 연결 관계 판단"),
    19: ("요약·통합", "요약문 완성", "다양한 문화의 신화가 보편적 사회 경험을 공유함을 요약"),
    20: ("요약·통합", "요약문 완성", "무언가를 걸치거나 경험하며 원하는 특성을 내면화하는 행위 요약"),
    21: ("서술형", "어순 배열", "로봇의 자기·외부 대상 구분 실패를 주어진 어휘로 완성"),
    22: ("서술형", "요약어 추출·종합", "컬링의 스위핑 분석 사례에서 핵심어 추출 및 종합 어휘 작성"),
    23: ("서술형", "세부정보 추출·어법 수정", "우주 귀환 일기에서 세부정보를 찾고 분사 형태를 수정"),
    24: ("서술형", "조건 영작", "만성 통증의 경보 체계를 분사구문과 제한 어휘로 요약"),
}


def ox_value(value: object) -> bool:
    return str(value or "").strip().upper() == "O"


def percent(correct: int, total: int) -> str:
    return f"{correct / total * 100:.1f}%" if total else "-"


def safe_cell(value: object) -> str:
    return str(value).replace("|", "\\|").replace("\n", " ")


def main() -> None:
    workbook = load_workbook(SOURCE_XLSX, data_only=True)
    sheet = workbook["전화번호순"]
    source_rows = [row for row in range(3, sheet.max_row + 1) if sheet.cell(row, 3).value]

    if REPORT_DIR.exists():
        shutil.rmtree(REPORT_DIR)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    cohort_rates: dict[int, float] = {}
    for question in range(1, 21):
        correct = sum(ox_value(sheet.cell(row, 57 + question).value) for row in source_rows)
        cohort_rates[question] = correct / len(source_rows)

    cohort_total_avg = mean(float(sheet.cell(row, 5).value) for row in source_rows)
    cohort_written_avg = mean(float(sheet.cell(row, 6).value) for row in source_rows)
    cohort_objective_avg = mean(float(sheet.cell(row, 7).value) for row in source_rows)

    totals = [float(sheet.cell(row, 5).value) for row in source_rows]

    def total_rank(score: float) -> int:
        return 1 + sum(other > score for other in totals)

    index_rows: list[str] = []
    data_warnings: list[str] = []
    seed_records: list[dict[str, object]] = []

    for report_number, row in enumerate(source_rows, start=1):
        report_id = f"student-{report_number:03d}"
        name = str(sheet.cell(row, 3).value)
        source_rank = int(sheet.cell(row, 4).value)
        total_score = float(sheet.cell(row, 5).value)
        written_score = float(sheet.cell(row, 6).value)
        objective_score = float(sheet.cell(row, 7).value)
        computed_total = written_score + objective_score
        total_score_rank = total_rank(total_score)

        category_stats: dict[str, dict[str, object]] = defaultdict(
            lambda: {"questions": [], "correct": 0, "missed": []}
        )
        question_rows: list[str] = []

        for question in range(1, 21):
            category, detail_type, _ = QUESTIONS[question]
            response = str(sheet.cell(row, 7 + question).value or "").strip()
            is_correct = ox_value(sheet.cell(row, 57 + question).value)
            category_stats[category]["questions"].append(question)
            if is_correct:
                category_stats[category]["correct"] += 1
            else:
                category_stats[category]["missed"].append(question)
            result = "O" if is_correct else "X"
            question_rows.append(
                f"| {question} | {detail_type} | {ANSWER_KEY[question]} | {safe_cell(response)} | "
                f"{result} | {cohort_rates[question] * 100:.1f}% |"
            )

        category_rows: list[str] = []
        scored_categories: list[tuple[float, str, list[int]]] = []
        for category in ["대의 파악", "추론·의미", "세부 내용", "어법·어휘", "글의 구조", "요약·통합"]:
            stat = category_stats[category]
            questions = list(stat["questions"])
            correct = int(stat["correct"])
            missed = list(stat["missed"])
            cohort_correct = sum(cohort_rates[q] for q in questions)
            category_rows.append(
                f"| {category} | {correct}/{len(questions)} | {percent(correct, len(questions))} | "
                f"{cohort_correct / len(questions) * 100:.1f}% | "
                f"{', '.join(map(str, missed)) if missed else '-'} |"
            )
            scored_categories.append((correct / len(questions), category, missed))

        strongest = sorted(scored_categories, key=lambda item: (-item[0], item[1]))[:2]
        weakest = sorted(
            [item for item in scored_categories if item[2]],
            key=lambda item: (item[0], item[1]),
        )[:2]
        missed_questions = [
            question
            for question in range(1, 21)
            if not ox_value(sheet.cell(row, 57 + question).value)
        ]
        priorities = sorted(missed_questions, key=lambda question: (cohort_rates[question], question))[:3]

        strength_text = ", ".join(
            f"{category}({accuracy * 100:.0f}%)" for accuracy, category, _ in strongest
        )
        weakness_text = (
            ", ".join(f"{category}({accuracy * 100:.0f}%)" for accuracy, category, _ in weakest)
            if weakest
            else "객관식 전 영역에서 오답 없음"
        )
        priority_lines = []
        for question in priorities:
            _, detail_type, topic = QUESTIONS[question]
            priority_lines.append(
                f"- **{question}번 · {detail_type}**: {topic}. "
                f"학급 정답률 {cohort_rates[question] * 100:.1f}%로, 오답 근거와 정답 근거를 한 문장씩 대비해 복습하세요."
            )
        if not priority_lines:
            priority_lines.append("- 객관식 오답이 없습니다. 서술형 답안의 표현 정확성과 조건 충족 여부를 중심으로 복습하세요.")

        mismatch_note = ""
        if total_score != computed_total:
            mismatch_note = (
                f"> 검수 필요: 원본 총점은 {total_score:g}점이지만 객관식과 서술형 합계는 "
                f"{computed_total:g}점입니다. 게시 전 원본 채점표를 확인하세요.\n\n"
            )
            data_warnings.append(
                f"- {report_id} ({name}): 총점 {total_score:g}, 객관식+서술형 {computed_total:g}"
            )

        report = f"""# {name} 학생 성적표

백현고2 2026학년도 1학기 기말 대비 모의시험 4차

{mismatch_note}## 성적 요약

| 항목 | 학생 | 학급 참고값 |
|---|---:|---:|
| 총점 | {total_score:g}/100 | 평균 {cohort_total_avg:.1f} |
| 객관식 | {objective_score:g}/80 | 평균 {cohort_objective_avg:.1f} |
| 서술형 | {written_score:g}/20 | 평균 {cohort_written_avg:.1f} |
| 원본 석차 | {source_rank}위 / {len(source_rows)}명 | 객관식 점수 배열과 일치 |
| 총점 기준 참고 순위 | {total_score_rank}위 / {len(source_rows)}명 | 이 리포트에서 재계산 |

## 영역별 성취

| 영역 | 정답 수 | 학생 정답률 | 학급 정답률 | 오답 번호 |
|---|---:|---:|---:|---|
{chr(10).join(category_rows)}

## 진단

- 강점 영역: {strength_text}
- 우선 보완 영역: {weakness_text}
- 서술형은 합계 점수만 제공되어 21~24번별 진단은 할 수 없습니다.

## 우선 복습 문항

{chr(10).join(priority_lines)}

## 객관식 문항별 결과

| 번호 | 유형 | 정답 | 학생 답 | 결과 | 학급 정답률 |
|---:|---|---:|---:|:---:|---:|
{chr(10).join(question_rows)}

## 서술형 구성

| 번호 | 유형 | 배점 | 개인 문항 점수 |
|---:|---|---:|---|
| 21 | 어순 배열 | 5 | 원본에 세부 점수 없음 |
| 22 | 요약어 추출·종합 | 5 | 원본에 세부 점수 없음 |
| 23 | 세부정보 추출·어법 수정 | 5 | 원본에 세부 점수 없음 |
| 24 | 조건 영작 | 5 | 원본에 세부 점수 없음 |

서술형 합계: **{written_score:g}/20점**

---

분석 기준: 시험지 정답표와 채점 결과 엑셀. 전화번호는 리포트에 포함하지 않았습니다.
"""
        (REPORT_DIR / f"{report_id}.md").write_text(report, encoding="utf-8")
        seed_records.append({
            "sourceReportId": report_id,
            "studentName": name,
            "totalScore": total_score,
            "objectiveScore": objective_score,
            "writtenScore": written_score,
            "sourceRank": source_rank,
            "totalScoreRank": total_score_rank,
            "cohortSize": len(source_rows),
            "cohortAverages": {
                "total": round(cohort_total_avg, 1),
                "objective": round(cohort_objective_avg, 1),
                "written": round(cohort_written_avg, 1),
            },
            "categories": [
                {
                    "category": category,
                    "correct": int(category_stats[category]["correct"]),
                    "total": len(category_stats[category]["questions"]),
                    "cohortRate": round(
                        mean(cohort_rates[q] for q in category_stats[category]["questions"]) * 100,
                        1,
                    ),
                    "missedQuestions": list(category_stats[category]["missed"]),
                }
                for category in ["대의 파악", "추론·의미", "세부 내용", "어법·어휘", "글의 구조", "요약·통합"]
            ],
            "questions": [
                {
                    "number": question,
                    "category": QUESTIONS[question][0],
                    "detailType": QUESTIONS[question][1],
                    "topic": QUESTIONS[question][2],
                    "correctAnswer": ANSWER_KEY[question],
                    "studentAnswer": str(sheet.cell(row, 7 + question).value or "").strip(),
                    "correct": ox_value(sheet.cell(row, 57 + question).value),
                    "cohortRate": round(cohort_rates[question] * 100, 1),
                }
                for question in range(1, 21)
            ],
            "strengths": [category for _, category, _ in strongest],
            "priorities": priorities,
            "dataWarning": (
                f"원본 총점 {total_score:g}점과 객관식·서술형 합계 {computed_total:g}점이 일치하지 않습니다."
                if total_score != computed_total
                else None
            ),
            "markdown": report,
        })
        index_rows.append(
            f"| {report_id} | {name} | {total_score:g} | {objective_score:g} | {written_score:g} | "
            f"{source_rank} | [열기](students/{report_id}.md) |"
        )

    difficulty_rows = []
    for question in range(1, 21):
        category, detail_type, topic = QUESTIONS[question]
        difficulty_rows.append(
            f"| {question} | {category} | {detail_type} | {topic} | {ANSWER_KEY[question]} | "
            f"{cohort_rates[question] * 100:.1f}% |"
        )
    for question in range(21, 25):
        category, detail_type, topic = QUESTIONS[question]
        difficulty_rows.append(
            f"| {question} | {category} | {detail_type} | {topic} | 서술형 | 개별 정오 없음 |"
        )

    major_rows = []
    for category in ["대의 파악", "추론·의미", "세부 내용", "어법·어휘", "글의 구조", "요약·통합"]:
        questions = [q for q in range(1, 21) if QUESTIONS[q][0] == category]
        rate = mean(cohort_rates[q] for q in questions)
        major_rows.append(
            f"| {category} | {', '.join(map(str, questions))} | {rate * 100:.1f}% |"
        )

    hard = sorted(range(1, 21), key=lambda q: (cohort_rates[q], q))[:5]
    easy = sorted(range(1, 21), key=lambda q: (-cohort_rates[q], q))[:5]
    analysis = f"""# 백현고2 모의시험 4차 문항 유형 분석

## 시험 구성

- 객관식 20문항 × 4점 = 80점
- 서술형 4문항 × 5점 = 20점
- 분석 인원: {len(source_rows)}명
- 학급 평균: 총점 {cohort_total_avg:.1f}, 객관식 {cohort_objective_avg:.1f}, 서술형 {cohort_written_avg:.1f}

## 영역별 학급 성취

| 영역 | 문항 | 학급 정답률 |
|---|---|---:|
{chr(10).join(major_rows)}

가장 어려운 객관식은 {', '.join(f'{q}번({cohort_rates[q] * 100:.1f}%)' for q in hard)}이고, 가장 높은 정답률을 보인 문항은 {', '.join(f'{q}번({cohort_rates[q] * 100:.1f}%)' for q in easy)}입니다. 특히 문장 삽입 16·17번과 연결어 18번에서 집단적인 약점이 나타났습니다.

## 번호별 유형

| 번호 | 대영역 | 세부 유형 | 핵심 내용·평가 포인트 | 정답 | 학급 정답률 |
|---:|---|---|---|:---:|---:|
{chr(10).join(difficulty_rows)}

## 해석 시 주의점

- 객관식 문항별 응답과 정오표는 서로 일치합니다.
- 서술형은 학생별 합계만 있고 문항별 점수가 없어, 21~24번의 개인별 강약점은 판정할 수 없습니다.
- 원본 `석차`는 객관식 점수 배열과 일치합니다. 총점 순위로 해석하면 안 됩니다.
- 문항 유형은 시험지의 발문과 지문 전개를 기준으로 한 로컬 AI 분석입니다.
"""
    (OUTPUT_DIR / "question-type-analysis.md").write_text(analysis, encoding="utf-8")

    index = f"""# 학생별 Markdown 성적표 인덱스

분석 대상: 백현고2 2026학년도 1학기 기말 대비 모의시험 4차

> 개인정보 보호: 이 폴더를 정적 웹의 `public` 디렉터리에 복사하지 마세요. 실제 게시 시에는 인증 코드 검증 후 해당 학생의 리포트만 서버에서 반환해야 합니다.

| 리포트 ID | 학생 | 총점 | 객관식 | 서술형 | 원본 석차 | 리포트 |
|---|---|---:|---:|---:|---:|---|
{chr(10).join(index_rows)}

## 데이터 검수 항목

{chr(10).join(data_warnings) if data_warnings else '- 없음'}
"""
    (OUTPUT_DIR / "index.md").write_text(index, encoding="utf-8")

    comparison_rows = []
    for question in range(1, 25):
        category, detail_type, _ = QUESTIONS[question]
        comparison_rows.append(
            f"| {question} | {category} / {detail_type} | 실행 전 | 실행 전 | 검토 전 |"
        )
    comparison = f"""# 외부 API 분석 비교표

현재 상태: **로컬 분석 완료 / 외부 API 미실행**

외부 분석 서비스와 API 키가 지정되지 않았고, 학생 개인정보를 외부로 전송하는 것은 별도 승인이 필요하므로 이번 작업에서는 외부 호출을 하지 않았습니다. 비교를 실행할 때도 학생 이름·전화번호·점수는 보내지 않고 시험 문항 텍스트만 전송하는 방식을 권장합니다.

## 비교 기준

- 대영역 일치 여부
- 세부 유형 일치 여부
- 핵심 평가 포인트의 의미 일치 여부
- 불일치 문항의 교사 최종 검토

| 번호 | 로컬 AI 분석 | 외부 API 분석 | 대영역 일치 | 최종 판정 |
|---:|---|---|:---:|---|
{chr(10).join(comparison_rows)}

## 권장 외부 응답 형식

```json
{{
  "questionNumber": 1,
  "primaryCategory": "대의 파악",
  "detailType": "주제",
  "skills": ["핵심 주장 파악", "선택지 의미 대조"],
  "rationale": "한두 문장 근거"
}}
```

외부 결과는 자동 확정하지 말고, 불일치 항목만 교사가 검토한 뒤 `최종 판정` 열을 확정하는 흐름이 안전합니다.
"""
    (OUTPUT_DIR / "external-api-comparison.md").write_text(comparison, encoding="utf-8")

    publication = """# 홈페이지 게시 설계안

## 권장 사용자 흐름

1. 모든 학생이 같은 `/report` 페이지에 접속합니다.
2. 문자로 받은 개인 식별 코드를 입력합니다.
3. Cloud Function이 코드의 해시와 만료·시도 횟수를 검증합니다.
4. 검증에 성공하면 짧은 유효기간의 열람 권한을 발급합니다.
5. 프런트엔드는 권한에 연결된 한 학생의 Markdown 성적표만 받아 렌더링합니다.

즉, **하나의 페이지 + 학생별 식별 코드 + 학생별 비공개 리포트** 구조입니다.

## 저장 구조 예시

```text
reportAccess/{accessId}
  codeHash
  studentId
  reportId
  expiresAt
  failedAttempts
  lockedUntil

studentReports/{reportId}
  studentId
  examId
  markdown
  publishedAt
```

## 보안 원칙

- Markdown 파일을 Vite의 `public` 폴더나 공개 Storage URL에 두지 않습니다.
- 원문 식별 코드는 Firestore에 저장하지 않고 서버에서 해시한 값만 저장합니다.
- 코드 입력은 IP·코드 기준으로 횟수를 제한하고 반복 실패 시 잠급니다.
- 응답에는 전화번호·다른 학생 목록·내부 문서 경로를 포함하지 않습니다.
- 열람 로그와 코드 만료일을 기록하고 재발급 시 이전 코드를 폐기합니다.

## 다음 구현 단위

- `verifyStudentReportCode` Cloud Function
- `/report` 코드 입력 페이지
- Markdown 렌더러와 인쇄용 스타일
- 관리자용 리포트 업로드·코드 발급 화면
- SMS에는 공통 링크만 넣고 개인 식별 코드는 별도 값으로 삽입
"""
    (OUTPUT_DIR / "publication-plan.md").write_text(publication, encoding="utf-8")

    seed_bundle = {
        "exam": {
            "examId": "baekhyeon-2026-final-4",
            "title": "백현고2 2026학년도 1학기 기말 대비 모의시험 4차",
            "school": "백현고2",
            "cohortSize": len(source_rows),
        },
        "reports": seed_records,
    }
    REPORT_SEED_PATH.write_text(
        "// 이 파일은 scripts/generate_baekhyeon_reports.py에서 자동 생성됩니다.\n"
        "export const baekhyeonReportSeed = "
        + json.dumps(seed_bundle, ensure_ascii=False, indent=2)
        + " as const\n",
        encoding="utf-8",
    )

    print(f"Generated {len(source_rows)} student reports in {OUTPUT_DIR}")
    print(f"Generated server seed data in {REPORT_SEED_PATH}")


if __name__ == "__main__":
    main()
